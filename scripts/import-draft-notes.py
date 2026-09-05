#!/usr/bin/env python3
"""
Import the scribe's draft notes out of Scott's workbook.

    python3 scripts/import-draft-notes.py ~/Downloads/'2026 Fantasy Football.xlsx'

Scott Johnston keeps the notes every year. He sits at the table with the
workbook open and writes down what people say as they say it, which makes his
`SCJ COMMENTS` column the only record that exists of the room — the app records
what was drafted and nothing whatsoever about the two hours around it.

WHAT THIS TAKES, AND WHAT IT DELIBERATELY LEAVES BEHIND
-------------------------------------------------------
Only the commentary. His sheet also carries `Player`, `Position`, `Team`,
`League Member` and a `NOTES` column, and the app already knows all of it from
the board itself:

  * The player, position and NFL team are on the board, spelled correctly. His
    sheet has "Deeboo Samuel", "Xaiver Worthy", "Hunter [Fucking] Henrey" and
    seven more like them — he is transcribing a room, not maintaining a
    database, and that is the right priority for a scribe. Importing his names
    would put a second, wrong spelling of ten players into the app.
  * `NOTES` is almost entirely pick provenance — "from Kyle", "Scott's Keeper",
    "from Colin (via Witte)". The app derives all of that from the traded-pick
    ledger and the keeper overlay, which are reconciled against the
    commissioner's workbook. A hand-kept second copy of a thing already
    reconciled elsewhere is exactly the drift this repo keeps getting bitten by.

So the notes join onto the canonical board BY OVERALL PICK NUMBER and contribute
one field: what was said. Verified: all 141 of his rows that correspond to a live
pick join, and the 19 that do not are exactly the 19 keepers, which occupy board
slots nobody spoke over.

`sheetPlayer` is carried anyway, and never rendered. It is the audit trail — the
verify script joins the notes to the board and checks his player against the
board's, so if he sends a re-cut sheet with a row inserted, the resulting
one-off misalignment is a loud failure rather than sixty quotes silently
attached to the wrong picks.
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is needed to read the workbook: pip3 install openpyxl")

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "data" / "draft-notes-2026.json"
SHEET = "Draft Board"
SEASON = 2026

# The scribe's own column, by its heading in his workbook.
COMMENT_HEADING = "SCJ COMMENTS"
NOTES_HEADING = "NOTES"

SCRIBE = {"shortName": "Scott", "fullName": "Scott Johnston"}

"""
Rows where the transcription is malformed and the fix is not a judgement call.

The renderer pairs quote marks to work out who said what, so a row with an odd
number of them would silently swallow the next speaker's line as an attribution.
Rather than guess in the renderer, the two known bad rows are corrected here
with the reason written down, and ANY OTHER malformed row is a hard failure — see
`check_quotes`. That way a re-cut sheet with a new typo stops the import instead
of quietly producing a garbled quote.

Keyed by overall pick. The value replaces the cell verbatim.
"""
CORRECTIONS = {
    # Missing the closing quote after "seriously?", which made the waitress's
    # line swallow Colin's reply as its attribution. The following
    # '"I told him he's a fucking nerd" Waitress' fixes who was speaking.
    118: (
        '"You guys take this shit so seriously?" Waitress '
        '"I BUILT THIS APP!" Colin to Waitress '
        '"I told him he\'s a fucking nerd" Waitress '
        '"The season doesn\'t start for 2 more weeks. He\'s going to rosterbate '
        'a lot. He\'s going to be dry" Stefan'
    ),
    # Missing the space between Kyle's attribution and Greg's opening quote.
    # Pairs correctly either way; fixed so the stored text reads as written.
    146: (
        '"That\'s my starter. Top 5 perennial TE in the 15th round" Kyle '
        '"Mark Andrews is 38 years old." Greg '
        '"No he isn\'t….is he?" Kyle'
    ),
}

"""
A quote Scott filed in the NOTES column instead of the commentary column.

Only one exists and it is unambiguously commentary — Witte drafting Ja'Marr
Chase to a cry of "Canada boy". The rest of NOTES is provenance this importer
drops on purpose, so rather than widen the rule, the single exception is named.
"""
NOTES_AS_COMMENTARY = {2}


def check_quotes(pick: int, text: str) -> None:
    """Refuse a row the renderer cannot pair up."""
    if text.count('"') % 2 == 0:
        return
    sys.exit(
        f"Pick {pick} has an odd number of quote marks and cannot be parsed:\n"
        f"  {text}\n"
        f"Add a correction to CORRECTIONS in {Path(__file__).name} with a note "
        f"on what was meant. Do not let the renderer guess."
    )


def main() -> None:
    if len(sys.argv) != 2:
        sys.exit(f"usage: python3 {Path(__file__).name} <path to workbook.xlsx>")

    book = Path(sys.argv[1]).expanduser()
    if not book.exists():
        sys.exit(f"No workbook at {book}")

    workbook = openpyxl.load_workbook(book, data_only=True)
    if SHEET not in workbook.sheetnames:
        sys.exit(f"{book.name} has no {SHEET!r} sheet. Found: {workbook.sheetnames}")

    sheet = workbook[SHEET]
    rows = sheet.iter_rows(values_only=True)
    heading = [str(c).strip() if c is not None else "" for c in next(rows)]

    for wanted in ("Round", "Pick", "Player", COMMENT_HEADING, NOTES_HEADING):
        if wanted not in heading:
            sys.exit(f"{SHEET!r} has no {wanted!r} column. Found: {heading}")

    col = {name: i for i, name in enumerate(heading)}
    notes = []
    seen = set()

    for row in rows:
        pick = row[col["Pick"]]
        if pick is None:
            continue
        pick = int(pick)
        if pick in seen:
            sys.exit(f"Pick {pick} appears twice in {SHEET!r}.")
        seen.add(pick)

        said = str(row[col[COMMENT_HEADING]] or "").strip()
        if not said and pick in NOTES_AS_COMMENTARY:
            said = str(row[col[NOTES_HEADING]] or "").strip()
        if pick in CORRECTIONS:
            said = CORRECTIONS[pick]
        if not said:
            continue

        # Collapse the runs of whitespace a spreadsheet cell picks up. The text
        # is otherwise verbatim: his spelling, his punctuation, his ellipses.
        said = re.sub(r"\s+", " ", said).strip()
        check_quotes(pick, said)

        notes.append(
            {
                "overallPick": pick,
                "round": int(row[col["Round"]]) if row[col["Round"]] else None,
                "sheetPlayer": str(row[col["Player"]] or "").strip() or None,
                "said": said,
            }
        )

    notes.sort(key=lambda n: n["overallPick"])

    OUT.write_text(
        json.dumps(
            {
                "note": (
                    "The scribe's record of the 2026 draft room, imported from the "
                    "SCJ COMMENTS column of Scott Johnston's workbook by "
                    "scripts/import-draft-notes.py. Joined onto the board by "
                    "overallPick; the board supplies every player name, because "
                    "the sheet is a transcript of a room and not a player "
                    "database. sheetPlayer is the audit trail for that join and "
                    "is never rendered. Do not hand-edit: re-run the importer."
                ),
                "season": SEASON,
                "scribe": SCRIBE,
                "source": book.name,
                "commentedPicks": len(notes),
                "notes": notes,
            },
            indent=1,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )

    print(f"wrote {len(notes)} commented picks to {OUT.relative_to(REPO)}")


if __name__ == "__main__":
    main()
